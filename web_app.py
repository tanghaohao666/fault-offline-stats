import os
import time
import tempfile
import requests
import pandas as pd
from pathlib import Path
from datetime import datetime
from flask import Flask, send_file, jsonify
import threading

app = Flask(__name__)

@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    return response

BASE_URL = "https://web.mdxev.com"
LOGIN_URL = f"{BASE_URL}/api/Account/PreLogin"
FINALIZE_URL = f"{BASE_URL}/api/Account/FinalizeLogin"
EXPORT_URL = f"{BASE_URL}/api/merchant/BigScreen/ExportChargeStateRecord"

USERNAME = os.environ.get("API_USERNAME", "mdx_yys_thys")
PASSWORD = os.environ.get("API_PASSWORD", "mdx123456")

ENTERPRISES = [
    "四川太和华坤新能源科技有限公司",
    "四川太和驭驷新能源科技有限公司",
    "四川谢尔顿新能源科技有限公司",
]

EXCLUDE_SITES = [
    "九鼎山·太子岭滑雪场二期直流充电站",
    "金科·集美天宸25栋充电站",
    "九鼎山·太子岭滑雪场二期交流充电站",
    "九鼎山·太子岭滑雪场充电站",
]

BASE_DIR = Path(__file__).parent
IS_CLOUD = os.environ.get("IS_CLOUD", "").lower() in ("1", "true", "yes")
if IS_CLOUD:
    DOWNLOAD_DIR = Path(tempfile.gettempdir()) / "fault_offline_downloads"
    OUTPUT_DIR = Path(tempfile.gettempdir()) / "fault_offline_outputs"
else:
    DOWNLOAD_DIR = BASE_DIR / "downloads"
    OUTPUT_DIR = BASE_DIR
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

task_status = {
    "running": False,
    "progress": "",
    "logs": [],
    "file_path": None,
    "error": None,
}
task_lock = threading.Lock()


def _log(msg, important=True):
    if important:
        task_status["logs"].append(msg)
        task_status["progress"] = msg


def _calc_width(text):
    w = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            w += 2
        else:
            w += 1
    return w


def _write_excel_with_autosize(df, ts):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(df.columns) + 1):
        max_w = _calc_width(df.columns[col_idx - 1])
        for row_idx in range(2, len(df) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                w = _calc_width(v)
                if w > max_w:
                    max_w = w
        adjusted = min(max(max_w + 2, 8), 50)
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = adjusted

    output_file = OUTPUT_DIR / f"{ts}故障离线站点统计.xlsx"
    try:
        wb.save(output_file)
    except PermissionError:
        output_file = OUTPUT_DIR / f"{ts}故障离线站点统计_{datetime.now().second}秒.xlsx"
        wb.save(output_file)

    return output_file


def _do_export():
    task_status["logs"] = []
    task_status["file_path"] = None
    task_status["error"] = None

    s = requests.Session()
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Origin": "https://merchant.mdxev.com",
        "Referer": "https://merchant.mdxev.com/",
    })

    _log("登录中...", important=False)
    r = s.post(LOGIN_URL, json={"userName": USERNAME, "password": PASSWORD})
    data = r.json()
    if not data.get("data") or not data["data"].get("tempToken"):
        raise RuntimeError(f"登录失败: {data}")
    temp_token = data["data"]["tempToken"]
    _log("登录成功", important=False)

    for f in DOWNLOAD_DIR.glob("*.xlsx"):
        f.unlink()

    for idx, enterprise in enumerate(ENTERPRISES):
        _log(f"处理企业 [{idx + 1}/{len(ENTERPRISES)}]: {enterprise}")

        r2 = s.post(LOGIN_URL, json={"userName": USERNAME, "password": PASSWORD})
        data2 = r2.json()
        temp_token = data2["data"]["tempToken"]
        enterprises_list = data2["data"]["enterprises"]

        ent_id = None
        for e in enterprises_list:
            if enterprise in e["enterpriseName"]:
                ent_id = e["enterpriseId"]
                break

        if not ent_id:
            _log(f"  错误: 未找到企业 {enterprise}")
            continue

        r3 = s.post(FINALIZE_URL, json={"tempToken": temp_token, "enterpriseId": ent_id})
        final_data = r3.json()
        if not final_data.get("data") or not final_data["data"].get("accessToken"):
            _log(f"  错误: FinalizeLogin 失败 - {final_data}")
            continue
        access_token = final_data["data"]["accessToken"]
        s.headers["Authorization"] = f"Bearer {access_token}"
        _log(f"已切换到: {final_data['data'].get('enterpriseName', 'N/A')}", important=False)

        for state_type, state_name in [(3, "故障"), (2, "离线")]:
            _log(f"导出{state_name}数据...", important=False)
            params = {"stateType": state_type, "chargeType": ""}
            r4 = s.post(EXPORT_URL, json=params, stream=True)

            content_type = r4.headers.get("content-type", "")
            if r4.status_code == 200 and "json" not in content_type and "text" not in content_type:
                ent_short = enterprise.replace("四川太和", "").replace("四川谢尔顿", "谢尔顿")
                filename = f"{ent_short}_{state_name}_明细.xlsx"
                filepath = DOWNLOAD_DIR / filename

                with open(filepath, "wb") as f:
                    for chunk in r4.iter_content(chunk_size=8192):
                        f.write(chunk)

                if filepath.stat().st_size > 100:
                    _log(f"已保存: {filename}")
                else:
                    filepath.unlink()
                    _log(f"文件太小，跳过: {filename}", important=False)
            else:
                _log(f"请求失败: HTTP {r4.status_code}")

            time.sleep(1)

    _log("合并数据中...", important=False)
    all_fault_dfs = []
    all_offline_dfs = []

    files = sorted(DOWNLOAD_DIR.glob("*.xlsx"))
    _log(f"发现 {len(files)} 个数据文件", important=False)

    for f in files:
        try:
            df = pd.read_excel(f)
            if "故障" in f.name:
                all_fault_dfs.append(df)
            elif "离线" in f.name:
                all_offline_dfs.append(df)
        except Exception as e:
            _log(f"读取失败: {f.name}", important=False)

    if not all_fault_dfs and not all_offline_dfs:
        raise RuntimeError("没有数据文件！")

    fault_df = pd.concat(all_fault_dfs, ignore_index=True) if all_fault_dfs else pd.DataFrame()
    offline_df = pd.concat(all_offline_dfs, ignore_index=True) if all_offline_dfs else pd.DataFrame()

    common_cols = ["运营商", "站点名称", "枪号", "时间", "生产商", "区域"]

    if not fault_df.empty and all(c in fault_df.columns for c in common_cols):
        fault_data = fault_df[common_cols + ["故障原因"]].copy()
        fault_data["状态"] = "故障"
        fault_data = fault_data[common_cols + ["状态", "故障原因"]]
    else:
        fault_data = pd.DataFrame(columns=common_cols + ["状态", "故障原因"])

    if not offline_df.empty and all(c in offline_df.columns for c in common_cols):
        offline_data = offline_df[common_cols].copy()
        offline_data["状态"] = "离线"
        offline_data["故障原因"] = ""
        offline_data = offline_data[common_cols + ["状态", "故障原因"]]
    else:
        offline_data = pd.DataFrame(columns=common_cols + ["状态", "故障原因"])

    merged = pd.concat([fault_data, offline_data], ignore_index=True)

    if "时间" in merged.columns:
        merged = merged.dropna(subset=["时间"])

    if "区域" in merged.columns:
        mask_dazhou = merged["区域"].astype(str).str.contains("四川省达州市", na=False)
        merged = merged[~mask_dazhou]

    mask_sites = merged["站点名称"].isin(EXCLUDE_SITES)
    merged = merged[~mask_sites]

    ts = datetime.now().strftime(f"{datetime.now().month}月{datetime.now().day}日{datetime.now().hour}时{datetime.now().minute}分")
    output_file = _write_excel_with_autosize(merged, ts)

    _log(f"\n完成! 共 {len(merged)} 条记录")
    _log(f"文件: {output_file.name}")

    task_status["file_path"] = str(output_file)


@app.route("/")
def index():
    return send_file(str(BASE_DIR / "index.html"))


@app.route("/api/start", methods=["POST"])
def start_export():
    with task_lock:
        if task_status["running"]:
            return jsonify({"status": "already_running"}), 409
        task_status["running"] = True

    def worker():
        try:
            _do_export()
        except Exception as e:
            task_status["error"] = str(e)
            task_status["logs"].append(f"错误: {e}")
        finally:
            task_status["running"] = False

    t = threading.Thread(target=worker, daemon=True)
    t.start()
    return jsonify({"status": "started"})


@app.route("/api/status")
def get_status():
    return jsonify({
        "running": task_status["running"],
        "progress": task_status["progress"],
        "logs": task_status["logs"][-20:],
        "file_path": task_status["file_path"],
        "error": task_status["error"],
    })


@app.route("/api/download")
def download():
    if not task_status["file_path"]:
        return "文件不存在", 404
    filepath = Path(task_status["file_path"])
    if not filepath.exists():
        return "文件不存在", 404
    return send_file(str(filepath), as_attachment=True, download_name=filepath.name)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 50)
    print("  故障/离线站点统计 - Web版")
    print(f"  访问: http://127.0.0.1:{port}")
    print("=" * 50)
    app.run(host="0.0.0.0", port=port, debug=False)