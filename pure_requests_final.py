import json
import time
import requests
from pathlib import Path
import pandas as pd
from datetime import datetime

BASE_URL = "https://web.mdxev.com"
LOGIN_URL = f"{BASE_URL}/api/Account/PreLogin"
FINALIZE_URL = f"{BASE_URL}/api/Account/FinalizeLogin"
EXPORT_URL = f"{BASE_URL}/api/merchant/BigScreen/ExportChargeStateRecord"

USERNAME = "mdx_yys_thys"
PASSWORD = "mdx123456"

ENTERPRISES = [
    "四川太和华坤新能源科技有限公司",
    "四川太和驭驷新能源科技有限公司",
    "四川谢尔顿新能源科技有限公司",
]

EXCLUDE_SITES = [
    "九鼎山·太子岭滑雪场二期直流充电站",
    "金科·集美天宸25栋充电站",
    "九鼎山·太子岭滑雪场二期交流充电站",
    "九鼎山·太子岭滑雪场充电站",
]

BASE_DIR = Path(__file__).parent
DOWNLOAD_DIR = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

s = requests.Session()
s.headers.update({
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Origin": "https://merchant.mdxev.com",
    "Referer": "https://merchant.mdxev.com/",
})


def login():
    print("1. 登录...")
    r = s.post(LOGIN_URL, json={"userName": USERNAME, "password": PASSWORD})
    data = r.json()
    temp_token = data["data"]["tempToken"]
    print(f"  预登录成功, tempToken: {temp_token[:30]}...")
    return temp_token


def select_enterprise(temp_token, enterprise_name):
    print(f"\n2. 选择企业: {enterprise_name}")
    
    r = s.post(LOGIN_URL, json={"userName": USERNAME, "password": PASSWORD})
    data = r.json()
    enterprises = data["data"]["enterprises"]
    
    ent_id = None
    for e in enterprises:
        if enterprise_name in e["enterpriseName"]:
            ent_id = e["enterpriseId"]
            break
    
    if not ent_id:
        raise ValueError(f"未找到企业: {enterprise_name}")
    
    print(f"  企业ID: {ent_id}")
    
    r2 = s.post(FINALIZE_URL, json={"tempToken": temp_token, "enterpriseId": ent_id})
    final_data = r2.json()
    access_token = final_data["data"]["accessToken"]
    s.headers["Authorization"] = f"Bearer {access_token}"
    print(f"  accessToken: {access_token[:40]}...")
    return access_token


def switch_enterprise(enterprise_name):
    print(f"\n切换到企业: {enterprise_name}")
    
    r = s.post(LOGIN_URL, json={"userName": USERNAME, "password": PASSWORD})
    data = r.json()
    enterprises = data["data"]["enterprises"]
    temp_token = data["data"]["tempToken"]
    
    ent_id = None
    for e in enterprises:
        if enterprise_name in e["enterpriseName"]:
            ent_id = e["enterpriseId"]
            break
    
    if not ent_id:
        raise ValueError(f"未找到企业: {enterprise_name}")
    
    print(f"  企业ID: {ent_id}")
    
    r2 = s.post(FINALIZE_URL, json={"tempToken": temp_token, "enterpriseId": ent_id})
    final_data = r2.json()
    access_token = final_data["data"]["accessToken"]
    s.headers["Authorization"] = f"Bearer {access_token}"
    print(f"  accessToken: {access_token[:40]}...")
    print(f"  当前企业: {final_data['data'].get('enterpriseName', 'N/A')}")
    return access_token


def export_data(state_type, state_name, enterprise_name=""):
    print(f"\n  导出{state_name}数据 (stateType={state_type})...")
    
    params = {"stateType": state_type, "chargeType": ""}
    
    r = s.post(EXPORT_URL, json=params, stream=True)
    
    content_type = r.headers.get("content-type", "")
    content_disposition = r.headers.get("content-disposition", "")
    
    print(f"  响应状态: {r.status_code}")
    print(f"  Content-Type: {content_type}")
    print(f"  Content-Disposition: {content_disposition}")
    
    if r.status_code == 200:
        ent_short = enterprise_name.replace("四川太和", "").replace("四川谢尔顿", "谢尔顿")
        filename = f"{ent_short}_{state_name}_明细.xlsx"
        filepath = DOWNLOAD_DIR / filename
        
        if "json" in content_type or "text" in content_type:
            text = r.text
            print(f"  响应是JSON: {text[:500]}")
            return None
        
        with open(filepath, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                f.write(chunk)
        
        if filepath.stat().st_size > 100:
            print(f"  已保存: {filename} ({filepath.stat().st_size} bytes)")
            return filepath
        else:
            filepath.unlink()
            print(f"  文件太小，可能是空数据")
            return None
    else:
        print(f"  请求失败: {r.status_code}")
        print(f"  响应: {r.text[:500]}")
        return None


def merge_data():
    print("\n" + "=" * 50)
    print("合并数据...")
    
    all_fault_dfs = []
    all_offline_dfs = []
    
    files = sorted(DOWNLOAD_DIR.glob("*.xlsx"))
    print(f"发现 {len(files)} 个文件:")
    
    for f in files:
        try:
            df = pd.read_excel(f)
            print(f"  {f.name}: {len(df)} 条记录, 列={list(df.columns)}")
            
            if "故障" in f.name:
                all_fault_dfs.append(df)
            elif "离线" in f.name:
                all_offline_dfs.append(df)
        except Exception as e:
            print(f"  读取失败: {f.name} - {e}")
    
    if not all_fault_dfs and not all_offline_dfs:
        print("没有数据文件！")
        return
    
    fault_df = pd.concat(all_fault_dfs, ignore_index=True) if all_fault_dfs else pd.DataFrame()
    offline_df = pd.concat(all_offline_dfs, ignore_index=True) if all_offline_dfs else pd.DataFrame()
    
    common_cols = ["运营商", "站点名称", "枪号", "时间", "生产商", "区域"]
    
    print(f"\n故障数据: {len(fault_df)} 条, 列={list(fault_df.columns)}")
    print(f"离线数据: {len(offline_df)} 条, 列={list(offline_df.columns)}")
    
    # 处理故障数据
    if not fault_df.empty and all(c in fault_df.columns for c in common_cols):
        fault_data = fault_df[common_cols + ["故障原因"]].copy()
        fault_data["状态"] = "故障"
        fault_data = fault_data[common_cols + ["状态", "故障原因"]]
        print(f"  故障数据处理后: {len(fault_data)} 条")
    else:
        if not fault_df.empty:
            missing = [c for c in common_cols if c not in fault_df.columns]
            print(f"  警告: 故障数据缺少列: {missing}")
            print(f"  实际列: {list(fault_df.columns)}")
        fault_data = pd.DataFrame(columns=common_cols + ["状态", "故障原因"])
    
    # 处理离线数据
    if not offline_df.empty and all(c in offline_df.columns for c in common_cols):
        offline_data = offline_df[common_cols].copy()
        offline_data["状态"] = "离线"
        offline_data["故障原因"] = ""
        offline_data = offline_data[common_cols + ["状态", "故障原因"]]
        print(f"  离线数据处理后: {len(offline_data)} 条")
    else:
        if not offline_df.empty:
            missing = [c for c in common_cols if c not in offline_df.columns]
            print(f"  警告: 离线数据缺少列: {missing}")
            print(f"  实际列: {list(offline_df.columns)}")
        offline_data = pd.DataFrame(columns=common_cols + ["状态", "故障原因"])
    
    merged = pd.concat([fault_data, offline_data], ignore_index=True)
    before_total = len(merged)
    print(f"\n合并后总行数: {before_total}")
    
    # 步骤1: 排除"时间"为空的行
    if "时间" in merged.columns:
        merged = merged.dropna(subset=["时间"])
        after_time = len(merged)
        print(f"  步骤1: 排除时间为空的行 -> {after_time} 条 (删除 {before_total - after_time} 条)")
        before_total = after_time
    
    # 步骤2: 排除区域在"四川省达州市"的行
    if "区域" in merged.columns:
        mask_dazhou = merged["区域"].astype(str).str.contains("四川省达州市", na=False)
        merged = merged[~mask_dazhou]
        after_region = len(merged)
        print(f"  步骤2: 排除四川省达州市的行 -> {after_region} 条 (删除 {before_total - after_region} 条)")
        before_total = after_region
    
    # 步骤3: 排除指定站点
    mask_sites = merged["站点名称"].isin(EXCLUDE_SITES)
    merged = merged[~mask_sites]
    after_sites = len(merged)
    print(f"  步骤3: 排除指定站点 -> {after_sites} 条 (删除 {before_total - after_sites} 条)")
    
    ts = datetime.now().strftime(f"{datetime.now().month}月{datetime.now().day}日{datetime.now().hour}时{datetime.now().minute}分")
    output_file = _write_excel_with_autosize(merged, ts)
    
    print(f"\n最终结果: {len(merged)} 条")
    print(f"已生成: {output_file}")
    
    return output_file


def _calc_width(text):
    w = 0
    for ch in str(text):
        if '\u4e00' <= ch <= '\u9fff' or '\u3000' <= ch <= '\u303f' or '\uff00' <= ch <= '\uffef':
            w += 2
        else:
            w += 1
    return w


def _write_excel_with_autosize(df, ts):
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    for col_idx in range(1, len(df.columns) + 1):
        max_w = _calc_width(df.columns[col_idx - 1])
        for row_idx in range(2, len(df) + 2):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                w = _calc_width(v)
                if w > max_w:
                    max_w = w
        adjusted = min(max(max_w + 2, 8), 50)
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = adjusted
    
    output_file = BASE_DIR / f"{ts}故障离线站点统计.xlsx"
    try:
        wb.save(output_file)
    except PermissionError:
        output_file = BASE_DIR / f"{ts}故障离线站点统计_{datetime.now().second}秒.xlsx"
        wb.save(output_file)
    
    print(f"  列宽已自动调整 ({len(df.columns)} 列)")
    return output_file


def main():
    print("=" * 50)
    print("  满电行 - 故障/离线数据下载工具 (纯requests版)")
    print("=" * 50)
    
    # 清空下载目录
    for f in DOWNLOAD_DIR.glob("*.xlsx"):
        f.unlink()
    
    # 登录
    temp_token = login()
    
    # 第一个企业
    first_enterprise = ENTERPRISES[0]
    print(f"\n{'='*50}")
    print(f"处理: {first_enterprise}")
    print("=" * 50)
    
    access_token = select_enterprise(temp_token, first_enterprise)
    
    # 导出故障 (stateType=3)
    export_data(3, "故障", first_enterprise)
    time.sleep(1)
    
    # 导出离线 (stateType=2)
    export_data(2, "离线", first_enterprise)
    time.sleep(1)
    
    # 处理其他企业
    for enterprise in ENTERPRISES[1:]:
        print(f"\n{'='*50}")
        print(f"处理: {enterprise}")
        print("=" * 50)
        
        switch_enterprise(enterprise)
        time.sleep(1)
        
        export_data(3, "故障", enterprise)
        time.sleep(1)
        export_data(2, "离线", enterprise)
        time.sleep(1)
    
    # 合并
    merge_data()
    
    print("\n" + "=" * 50)
    print("全部完成!")


if __name__ == "__main__":
    main()